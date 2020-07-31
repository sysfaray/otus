# -*- coding: utf-8 -*-
# ----------------------------------------------------------------------
# Copyright (C) 2020 Maksim Mikhnenko
# ----------------------------------------------------------------------

# Python modules
import openpyxl
import logging
from collections import namedtuple

# Django modules
from django.contrib import admin
from django.db import models
from django import forms
from django.conf.urls import url
from django.shortcuts import render
from django.http import HttpResponseRedirect

# Project modules
from .models import Car
from brand.models import Brand, BrandModel

logger = logging.getLogger("django.request")

COL_MAP_EXCEL = {
    0: ("brand", str),
    1: ("model", str),
    2: ("capacity", float),
    3: ("fuel", str),
    4: ("etype", str),
    5: ("start_year", int),
    6: ("end_year", str),
}
Problem = namedtuple("Problem", ["line", "message"])


class ImportForm(forms.Form):
    file = forms.FileField()


@admin.register(Car)
class CarAdmin(admin.ModelAdmin):
    change_list_template = "admin/monitor_change_list.html"

    list_display = (
        "active",
        "model",
        "fuel",
        "etype",
        "capacity",
        "start_year",
        "end_year",
        "now",
    )
    list_display_links = ["model"]
    fields = [
        "active",
        "model",
        ("fuel", "etype"),
        "capacity",
        ("start_year", "end_year", "now"),
    ]
    list_filter = ["model__brand__name", "fuel", "etype"]
    search_fields = ["model__name"]
    formfield_overrides = {
        models.CharField: {"widget": forms.TextInput(attrs={"size": "20"})}
    }

    problems = []
    enabled = set()
    result = {"brand": 0, "model": 0, "cars": 0}

    def get_urls(self):
        urls = super(CarAdmin, self).get_urls()
        custom_urls = [
            url("^import/$", self.import_excel, name="import_excel"),
        ]
        return custom_urls + urls

    def import_excel(self, request):
        if request.method == "POST":
            file = request.FILES["file"]
            excel_file = openpyxl.load_workbook(file)
            for n, r in self.iter_data(excel_file):
                result = self.add_car(r)
            self.message_user(
                request,
                "Your excl file has been imported\n"
                "Create Brand: %s, Model: %s, Cars:%s"
                % (result["brand"], result["model"], result["cars"]),
            )
            self.result = {"brand": 0, "model": 0, "cars": 0}
            return HttpResponseRedirect("..")
        form = ImportForm()
        payload = {"form": form}
        return render(request, "admin/excel_form.html", payload)

    def iter_excel(self, excel_file):
        for sn in excel_file.sheetnames:
            for n, row in enumerate(excel_file[sn].iter_rows()):
                value = [r.value for r in row]
                try:
                    self.enabled.add((value[0], value[1]))
                except IndexError:
                    logger.warning("Index error on: ", n, value)
            continue
        for n, row in enumerate(excel_file[sn].iter_rows()):
            yield [r.value if r.value else "" for r in row]

    def iter_data(self, excel_file):
        row = self.iter_excel(excel_file)
        r_map = COL_MAP_EXCEL

        for n, row in enumerate(row):
            if n == 0:
                continue
            has_problem = False
            r = {}
            for cn in r_map:
                fn, tc = r_map[cn]
                v = row[cn]
                if not isinstance(v, int):
                    v = v.strip()
                if not v:
                    v = None
                if v:
                    try:
                        v = tc(v)
                    except Exception as e:
                        logger.warning(e)
                        self.register_problem(
                            n, "Неверный формат поля %s: %s" % (fn, v)
                        )
                        has_problem = True
                r[fn] = v
            if not has_problem:
                yield n, r

    def register_problem(self, line, message):
        logger.info(line, message)
        self.problems += [Problem(line=line + 1, message=message)]

    def add_car(self, data):
        try:
            brand = Brand.objects.get(name=data["brand"])
        except Exception as e:
            logger.warning("Not Brand", e)
            brand = Brand.objects.create(name=data["brand"])
            self.result["brand"] = self.result["brand"] + 1

        try:
            model = BrandModel.objects.get(name=data["model"], brand=brand)
        except Exception as e:
            logger.warning("Not Brand Model", e)
            model = BrandModel.objects.create(name=data["model"], brand=brand)
            self.result["model"] = self.result["model"] + 1
        now = False
        if data["end_year"] == "Now":
            end = 0
            now = True
        else:
            end = data["end_year"]
        etype = data["etype"] if data["etype"] else ""
        if not Car.objects.filter(
            model=model,
            fuel=data["fuel"],
            etype=etype,
            capacity=data["capacity"],
            start_year=data["start_year"],
            end_year=end,
            now=now,
        ):
            try:
                Car.objects.create(
                    model=model,
                    fuel=data["fuel"],
                    etype=etype,
                    capacity=data["capacity"],
                    start_year=data["start_year"],
                    end_year=end,
                    now=now,
                )
                self.result["cars"] = self.result["cars"] + 1
            except Exception as e:
                logger.warning("Error:", e)
        else:
            logger.info("Сar %s in base %s %s" % (model, data["fuel"], etype))
        return self.result
