# -*- coding: utf-8 -*-
# ----------------------------------------------------------------------
# Copyright (C) 2020 Maksim Mikhnenko
# ----------------------------------------------------------------------

# Python modules
import openpyxl
import logging

# Django modules
from django.contrib import admin
from django.conf.urls import url
from django import forms
from django.shortcuts import render
from django.http import HttpResponseRedirect

# Project modules
from .models import Spares, SparesName, SparesAnalog
from car.models import Car
from brand.models import BrandModel

logger = logging.getLogger("django.request")


def duplicate_jorn(modeladmin, request, queryset):
    post_url = request.META["HTTP_REFERER"]

    for object in queryset:
        old = object.model.all()
        object.part_number = object.part_number + "-copy"
        object.id = None
        object.save()
        if old:
            for m in old:
                object.model.add(m)
                object.save()
    return HttpResponseRedirect(post_url)


duplicate_jorn.short_description = "Duplicate Selected"


class ImportForm(forms.Form):
    file = forms.FileField()


@admin.register(SparesName)
class SparesNameAdmin(admin.ModelAdmin):
    list_display = ("name", "active")


@admin.register(Spares)
class SparesAdmin(admin.ModelAdmin):
    change_list_template = "admin/monitor_change_list.html"

    list_display = ("name", "part_number", "count", "cost")
    list_display_links = ("name", "part_number")
    list_filter = [
        "model__model__brand__name",
        "model__model__name",
        "model__fuel",
        "model__capacity",
    ]
    search_fields = [
        "name__name",
        "part_number",
        "model__model__name",
        "model__model__brand__name",
    ]
    list_editable = ("cost", "count")
    filter_horizontal = ["model"]
    actions = [duplicate_jorn]

    problems = []
    enabled = set()

    def get_urls(self):
        urls = super(SparesAdmin, self).get_urls()
        custom_urls = [
            url("^import/$", self.import_excel, name="import_excel"),
        ]
        return custom_urls + urls

    def import_excel(self, request):
        if request.method == "POST":
            file = request.FILES["file"]
            excel_file = openpyxl.load_workbook(file)
            result = self.add_spares(self.iter_data(excel_file))

            self.message_user(
                request,
                "Your excel file has been imported" "Add Spares: %s" % result["spares"],
            )
            return HttpResponseRedirect("..")
        form = ImportForm()
        payload = {"form": form}
        return render(request, "admin/excel_form.html", payload)

    def iter_excel(self, excel_file):
        for sn in excel_file.sheetnames:
            for n, row in enumerate(excel_file[sn].iter_rows()):
                value = [r.value for r in row]
                try:
                    self.enabled.add((value[0], value[1]))
                except IndexError:
                    logger.error("Index error on: %s %s", n, value)
            continue
        for n, row in enumerate(excel_file[sn].iter_rows()):
            yield [r.value if r.value else "" for r in row]

    def iter_data_name(self, excel_file):
        result = {}
        row = self.iter_excel(excel_file)
        for n, row in enumerate(row):
            if n == 0:
                for m in row:
                    if not m:
                        continue
                    result[m.strip()] = row.index(m)
            else:
                continue
        return result

    def iter_data(self, excel_file):
        name_objects = self.iter_data_name(excel_file)
        row = self.iter_excel(excel_file)
        r = {}
        for n, row in enumerate(row):
            if n == 0 or n == 1:
                continue
            if row[0] and row[1]:
                logger.warning("Row error on: %s", row)
                continue
            if row[0] is None:
                logger.warning("Row error on: %s", row)
                continue

            for name, index in name_objects.items():
                if not row[index] and not row[index + 1] and not row[index + 2]:
                    continue
                if r.get(name):
                    r[name].append(
                        {
                            "name": row[1].strip(),
                            "part_number": row[index],
                            "price": row[index + 1],
                            "count": row[index + 2],
                        }
                    )
                else:
                    r[name] = [
                        {
                            "name": row[1].strip(),
                            "part_number": row[index],
                            "price": row[index + 1],
                            "count": row[index + 2],
                        }
                    ]
        return r

    def add_spares(self, data):
        res_s = 0
        for d, value in data.items():
            dd = d.split(",")
            now = False
            if len(dd) < 2:
                continue
            model = BrandModel.objects.get(name=dd[0])
            if dd[5].strip() == "Now":
                end = 0
                now = True
            else:
                end = dd[5]
            try:
                car = Car.objects.get(
                    model=model,
                    fuel=dd[2].strip(),
                    etype=dd[3].strip(),
                    capacity=dd[1].strip(),
                    start_year=dd[4].strip(),
                    end_year=end,
                    active=True,
                    now=now,
                )
            except Exception as e:
                logger.error("Error: %s", e)
                continue
            for sp in value:
                name = sp.get("name")
                part_number = (
                    sp.get("part_number") if sp.get("part_number") else name.lower()
                )

                cost = sp.get("price")
                sp_ok = Spares.objects.filter(part_number=part_number)
                if sp_ok:

                    sp_ok[0].model.add(car)
                    sp_ok[0].save()
                else:
                    spname = SparesName.objects.filter(name=name)
                    if spname:
                        res = Spares.objects.create(
                            name=spname[0], part_number=part_number, cost=cost, count=0
                        )
                        res.model.add(car)
                        res.save()
                        res_s = res_s + 1
                    else:
                        spname = SparesName.objects.create(name=name)
                        res = Spares.objects.create(
                            name=spname, part_number=part_number, cost=cost, count=0
                        )
                        res.model.add(car)
                        res.save()
                        res_s = res_s + 1
        return {"spares": res_s}


@admin.register(SparesAnalog)
class SparesAnalogAdmin(admin.ModelAdmin):
    change_list_template = "admin/monitor_change_list.html"

    list_display = ("name", "part_number", "count", "cost")
    list_display_links = ("name", "part_number")
    list_filter = [
        "model__model__brand__name",
        "model__model__name",
        "model__fuel",
        "model__capacity",
    ]
    search_fields = [
        "name__name",
        "part_number",
        "model__model__name",
        "model__model__brand__name",
    ]
    list_editable = ("cost", "count")
    filter_horizontal = ["model"]
    actions = [duplicate_jorn]

    problems = []
    enabled = set()

    def get_urls(self):
        urls = super(SparesAnalogAdmin, self).get_urls()
        custom_urls = [
            url("^import/$", self.import_excel, name="import_excel"),
        ]
        return custom_urls + urls

    def import_excel(self, request):
        if request.method == "POST":
            file = request.FILES["file"]
            excel_file = openpyxl.load_workbook(file)
            result = self.add_spares(self.iter_data(excel_file))

            self.message_user(
                request,
                "Your excl file has been imported<br>"
                "Add Spares: %s" % result["spares"],
            )
            return HttpResponseRedirect("..")
        form = ImportForm()
        payload = {"form": form}
        return render(request, "admin/excel_form.html", payload)

    def iter_excel(self, excel_file):
        for sn in excel_file.sheetnames:
            for n, row in enumerate(excel_file[sn].iter_rows()):
                value = [r.value for r in row]
                try:
                    self.enabled.add((value[0], value[1]))
                except IndexError:
                    logger.error("Index error on:  %s %s", n, value)
            continue
        for n, row in enumerate(excel_file[sn].iter_rows()):
            yield [r.value if r.value else "" for r in row]

    def iter_data_name(self, excel_file):
        result = {}
        row = self.iter_excel(excel_file)
        for n, row in enumerate(row):
            if n == 0:
                for m in row:
                    if not m:
                        continue
                    result[m.strip()] = row.index(m)
            else:
                continue
        return result

    def iter_data(self, excel_file):
        name_objects = self.iter_data_name(excel_file)
        row = self.iter_excel(excel_file)
        r = {}
        for n, row in enumerate(row):
            if n == 0 or n == 1:
                continue
            if row[0] is None:
                logger.error("Row error on: %s", row)
                continue

            for name, index in name_objects.items():
                if not row[index] and not row[index + 1] and not row[index + 2]:
                    continue
                if r.get(name):
                    r[name].append(
                        {
                            "name": row[0].strip(),
                            "part_number": row[index],
                            "price": row[index + 1],
                            "count": row[index + 2],
                        }
                    )
                else:
                    r[name] = [
                        {
                            "name": row[0].strip(),
                            "part_number": row[index],
                            "price": row[index + 1],
                            "count": row[index + 2],
                        }
                    ]
        return r

    def add_spares(self, data):
        res_s = 0
        for d, value in data.items():
            dd = d.split(",")
            now = False
            if len(dd) < 2:
                continue
            model = BrandModel.objects.get(name=dd[0])
            if dd[5].strip() == "Now":
                end = 0
                now = True
            else:
                end = dd[5]
            try:
                car = Car.objects.get(
                    model=model,
                    fuel=dd[2].strip(),
                    etype=dd[3].strip(),
                    capacity=dd[1].strip(),
                    start_year=dd[4].strip(),
                    end_year=end,
                    active=True,
                    now=now,
                )
            except Exception as e:
                logger.error("Error: %s", e)
                continue
            for sp in value:
                name = sp.get("name")
                part_number = (
                    sp.get("part_number") if sp.get("part_number") else name.lower()
                )
                cost = sp.get("price")
                sp_ok = Spares.objects.filter(part_number=part_number)
                if sp_ok:
                    sp_ok[0].model.add(car)
                    sp_ok[0].save()
                else:
                    spname = SparesName.objects.filter(name=name)
                    if spname:
                        res = Spares.objects.create(
                            name=spname[0], part_number=part_number, cost=cost, count=0
                        )
                        res.model.add(car)
                        res.save()
                        res_s = res_s + 1
                    else:
                        spname = SparesName.objects.create(name=name)
                        res = Spares.objects.create(
                            name=spname, part_number=part_number, cost=cost, count=0
                        )
                        res.model.add(car)
                        res.save()
                        res_s = res_s + 1
        return {"spares": res_s}
