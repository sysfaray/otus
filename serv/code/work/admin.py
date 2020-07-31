# -*- coding: utf-8 -*-
# ----------------------------------------------------------------------
# Copyright (C) 2020 Maksim Mikhnenko
# ----------------------------------------------------------------------

# Python modules
import openpyxl
import logging

# Django modules
from django.contrib import admin
from django.conf.urls import url
from django import forms
from django.shortcuts import render
from django.http import HttpResponseRedirect

# Project modules
from .models import Work, TypeWork, SparesAnalogCount, SparesOriginalCount
from spares.models import Spares
from car.models import Car
from brand.models import BrandModel

logger = logging.getLogger("django.request")


def duplicate_jorn(modeladmin, request, queryset):
    post_url = request.META["HTTP_REFERER"]

    change = False
    for object in queryset:
        old_os = object.original.all()
        old_oa = object.analog.all()
        old_c = object.cars.all()
        object.name = object.name + "-copy"
        object.id = None
        object.save()
        if old_os:
            for ms in old_os:
                object.analog.add(ms)
                change = True
        if old_oa:
            for ma in old_oa:
                object.original.add(ma)
                change = True
        if old_c:
            for c in old_c:
                object.cars.add(c)
                change = True
        if change:
            object.save()
    return HttpResponseRedirect(post_url)


duplicate_jorn.short_description = "Duplicate Selected"


class ImportForm(forms.Form):
    file = forms.FileField()


@admin.register(TypeWork)
class TypeWorkAdmin(admin.ModelAdmin):
    list_display = ("name", "active")


class SparesOriginalCountinline(admin.TabularInline):
    model = SparesOriginalCount
    extra = 1  # how many rows to show


class SparesAnalogCountCountinline(admin.TabularInline):
    model = SparesAnalogCount
    extra = 1  # how many rows to show


class WorkAdmin(admin.ModelAdmin):

    change_list_template = "admin/monitor_change_list.html"
    inlines = (SparesOriginalCountinline, SparesAnalogCountCountinline)
    filter_horizontal = ["cars"]
    list_display = ("name", "type_work", "work_ratio", "active")
    list_display_links = ("name", "type_work")
    list_filter = [
        "cars__model__brand__name",
        "cars__model__name",
        "cars__fuel",
        "cars__capacity",
    ]
    search_fields = [
        "name",
        "original__part_number",
        "analog__part_number",
        "cars__model__name",
        "cars__model__brand__name",
    ]
    list_editable = ("work_ratio",)
    actions = [duplicate_jorn]

    enabled = set()

    def get_urls(self):
        urls = super(WorkAdmin, self).get_urls()
        custom_urls = [
            url("^import/$", self.import_excel, name="import_excel"),
        ]
        return custom_urls + urls

    def import_excel(self, request):
        if request.method == "POST":
            file = request.FILES["file"]
            excel_file = openpyxl.load_workbook(file)
            result = self.add_spares(self.iter_data(excel_file))
            self.message_user(
                request,
                "Your excel file has been imported Add Works: %s, Spares not to Base: %s (%s)" % (result["spares"], result["not_to_base"], result["res_spares"]),
            )
            return HttpResponseRedirect("..")
        form = ImportForm()
        payload = {"form": form}
        return render(request, "admin/excel_form.html", payload)

    def iter_excel(self, excel_file):
        for sn in excel_file.sheetnames:
            for n, row in enumerate(excel_file[sn].iter_rows()):
                value = [r.value for r in row]
                try:
                    self.enabled.add((value[0], value[1]))
                except IndexError:
                    logger.error("Index error on: ", n, value)
            continue
        for n, row in enumerate(excel_file[sn].iter_rows()):
            yield [r.value if r.value else "" for r in row]

    def iter_data_name(self, excel_file):
        result = {}
        row = self.iter_excel(excel_file)
        for n, row in enumerate(row):
            if n == 0:
                for m in row:
                    if not m:
                        continue
                    result[m.strip()] = row.index(m)
            else:
                continue
        return result

    def iter_data(self, excel_file):
        name_objects = self.iter_data_name(excel_file)
        row = self.iter_excel(excel_file)
        r = {}
        for n, row in enumerate(row):
            if n == 0 or n == 1:
                continue
            if row[0] is None:
                logger.warning("Row error on: ", row)
                continue
            for name, index in name_objects.items():
                if not row[index] and not row[index + 1] and not row[index + 2]:
                    continue
                if row[0] != "" and row[1] != "":
                    wtype = row[0]
                    work = row[1]
                    continue

                if r.get(name):
                    if wtype in r[name] and work in r[name][wtype]:
                        r[name][wtype][work].append(
                            {
                                "name": row[1].strip(),
                                "part_number": row[index],
                                "price": row[index + 1],
                                "count": row[index + 2],
                            }
                        )
                    elif wtype in r[name] and work not in r[name][wtype]:
                        r[name][wtype][work] = [
                            {
                                "name": row[1].strip(),
                                "part_number": row[index],
                                "price": row[index + 1],
                                "count": row[index + 2],
                            }
                        ]
                    else:
                        r[name][wtype] = {
                            work: [
                                {
                                    "name": row[1].strip(),
                                    "part_number": row[index],
                                    "price": row[index + 1],
                                    "count": row[index + 2],
                                }
                            ]
                        }
                else:
                    r[name] = {}
                    if wtype in r[name] and work in r[name][wtype]:
                        r[name][wtype] = {
                            work: [
                                {
                                    "name": row[1].strip(),
                                    "part_number": row[index],
                                    "price": row[index + 1],
                                    "count": row[index + 2],
                                }
                            ]
                        }
                    else:
                        r[name][wtype] = {
                            work: [
                                {
                                    "name": row[1].strip(),
                                    "part_number": row[index],
                                    "price": row[index + 1],
                                    "count": row[index + 2],
                                }
                            ]
                        }
        return r

    def add_spares(self, data):
        res_w = 0
        not_res = []
        for d, value in data.items():
            dd = d.split(",")
            now = False
            if len(dd) < 2:
                continue
            model = BrandModel.objects.get(name=dd[0])
            if dd[5].strip() == "Now":
                end = 0
                now = True
            else:
                end = dd[5]
            try:
                car = Car.objects.get(
                    model=model,
                    fuel=dd[2].strip(),
                    etype=dd[3].strip(),
                    capacity=dd[1].strip(),
                    start_year=dd[4].strip(),
                    end_year=end,
                    active=True,
                    now=now,
                )
            except Exception as e:
                logger.error("Error:", e)
                continue
            for wt in value:
                try:
                    logger.debug("Search Type Work %s" % wt)
                    tw = TypeWork.objects.get(name=wt)
                except Exception as e:
                    logger.debug("Create Type Work %s" % wt)
                    tw = TypeWork.objects.create(name=wt)
                except Exception as e:
                    logger.error("Type Work WTF?:", e)
                check = False
                for wrk in value[wt]:
                    for ww in Work.objects.filter(name=wrk, type_work=tw):
                        for sp in value[wt][wrk]:
                            name = sp.get("name")
                            part_number = (
                                sp.get("part_number")
                                if sp.get("part_number")
                                else name.lower()
                            )
                            try:
                                res = Spares.objects.get(part_number=part_number)
                            except Exception as e:
                                logger.debug("Spares %s not in Base:", part_number)
                                not_res.append(part_number)
                                break
                            count = sp.get("count") if sp.get("count") != "" else 1
                            if SparesOriginalCount.objects.filter(
                                work__id=ww.id, name=res, count=count
                            ):
                                check = True
                            else:

                                check = False
                                break

                        if check:
                            ww.cars.add(car)
                            check = False
                        continue
                    else:
                        if SparesOriginalCount.objects.filter(
                            work__name=wrk, work__cars=car
                        ):

                            for sp in value[wt][wrk]:
                                name = sp.get("name")
                                part_number = (
                                    sp.get("part_number")
                                    if sp.get("part_number")
                                    else name.lower()
                                )
                                res = Spares.objects.get(part_number=part_number)
                                count = sp.get("count") if sp.get("count") != "" else 1
                                if not SparesOriginalCount.objects.filter(
                                    work__name=wrk,
                                    work__cars=car,
                                    name=res,
                                    count=count,
                                ):
                                    SparesOriginalCount.objects.create(
                                        work=SparesOriginalCount.objects.filter(
                                            work__name=wrk, work__cars=car
                                        )[0].work,
                                        name=res,
                                        count=count,
                                    )
                            continue
                        else:
                            for sp in value[wt][wrk]:
                                new_work = Work.objects.filter(
                                    name=wrk, type_work=tw, cars=car
                                )
                                name = sp.get("name")
                                part_number = (
                                    sp.get("part_number")
                                    if sp.get("part_number")
                                    else name.lower()
                                )
                                if part_number in not_res:
                                    continue
                                try:
                                    res = Spares.objects.get(part_number=part_number)
                                except Exception as e:

                                    logger.debug("Spares %s not in Base:", part_number)
                                    not_res.append(part_number)
                                    continue
                                count = sp.get("count") if sp.get("count") != "" else 1
                                if new_work:

                                    if not SparesOriginalCount.objects.filter(
                                        work=new_work[0], name=res, count=count
                                    ):

                                        SparesOriginalCount.objects.create(
                                            work=new_work[0], name=res, count=count
                                        )
                                        continue
                                else:

                                    new_work3 = Work.objects.create(
                                        name=wrk, type_work=tw
                                    )
                                    res_w = res_w + 1
                                    SparesOriginalCount.objects.create(
                                        work=new_work3, name=res, count=count
                                    )
                                    new_work3.cars.add(car)
                                    new_work3.save()
                                continue
        return {"spares": res_w, "not_to_base": len(set(not_res)), "res_spares": ",".join(q for q in set(not_res)) if not_res else ""}


admin.site.register(Work, WorkAdmin)
